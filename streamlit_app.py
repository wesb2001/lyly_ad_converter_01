import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
from openpyxl.styles import PatternFill, Font
import io
from auto_convert_excel import convert_excel_file

def main():
    st.set_page_config(
        page_title="LYLYL 광고 데이터 변환기",
        page_icon="📊",
        layout="centered"
    )

    st.title("LYLYL 광고 데이터 변환기")
    st.markdown("""
    ### 📌 사용 방법
    1. Meta 광고 관리자에서 다운로드한 엑셀 파일을 업로드하세요.
    2. 변환이 완료되면 자동으로 다운로드가 시작됩니다.
    
    ### 🔄 변환 내용
    - 컬럼명 변경 및 정리
    - 광고비 0인 행 제거
    - 후크 및 지속 지표 계산
    - 숫자 포맷 변경 (%, 소수점 등)
    - 조건부 서식 적용
        - ROAS: ≥3.0 파란색, ≥2.5 초록색, ≥1.0 주황색, <1.0 빨간색
        - CPC: <1000원 파란색, <1500원 초록색, <2000원 주황색, ≥2000원 빨간색
        - CVR: ≥7% 파란색, ≥5% 초록색, ≥3% 주황색, <3% 빨간색
        - CTR: ≥5% 파란색, ≥3% 초록색, ≥2% 주황색, <2% 빨간색
        - 후크: ≥40% 파란색, ≥30% 초록색, ≥20% 주황색, <20% 빨간색
        - 지속: ≥30% 파란색, ≥20% 초록색, ≥10% 주황색, <10% 빨간색
    """)

    uploaded_file = st.file_uploader("Excel 파일을 선택하세요", type=['xlsx', 'xls'])

    if uploaded_file is not None:
        try:
            # 임시 파일로 저장
            input_path = "temp_input.xlsx"
            with open(input_path, "wb") as f:
                f.write(uploaded_file.getvalue())

            # 데이터프레임 생성하여 날짜 정보 추출
            df = pd.read_excel(input_path)
            start_date = pd.to_datetime(df['보고 시작'].iloc[0]).strftime('%y%m%d')
            end_date = pd.to_datetime(df['보고 종료'].iloc[0]).strftime('%y%m%d')
            
            # 출력 파일명 생성
            output_filename = f"LYLYL_{start_date}_{end_date}_v01.xlsx"
            output_path = "temp_output.xlsx"

            # 파일 변환
            convert_excel_file(input_path, output_path)

            # 변환된 파일 다운로드
            with open(output_path, "rb") as f:
                bytes_data = f.read()
            
            st.download_button(
                label="변환된 파일 다운로드",
                data=bytes_data,
                file_name=output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # 임시 파일 삭제
            os.remove(input_path)
            os.remove(output_path)

            st.success("✅ 파일 변환이 완료되었습니다!")

        except Exception as e:
            st.error(f"❌ 오류가 발생했습니다: {str(e)}")
            st.error("파일 형식을 확인해주세요.")

if __name__ == "__main__":
    main() 