import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
from openpyxl.styles import PatternFill, Font
import glob

def get_next_version(base_filename):
    # 오늘 생성된 같은 날짜의 파일들 검색
    pattern = f"{base_filename}_v*.xlsx"
    existing_files = glob.glob(pattern)
    
    if not existing_files:
        return "v01"
    
    # 기존 파일들의 버전 번호 추출
    versions = []
    for file in existing_files:
        try:
            version = file.split("_")[-1].replace(".xlsx", "")  # v01, v02 등 추출
            version_num = int(version.replace("v", ""))
            versions.append(version_num)
        except:
            continue
    
    if not versions:
        return "v01"
    
    # 최대 버전 번호에 1을 더해서 반환
    next_version = max(versions) + 1
    return f"v{next_version:02d}"

def convert_excel_file(input_path: str, output_path: str):
    # 1. Load file
    df = pd.read_excel(input_path)

    # 2. Column mapping
    column_mapping = {
        "광고 이름": "제목",
        "광고 게재": "상태",
        "지출 금액 (KRW)": "광고비",
        "구매": "구매",
        "구매 전환값": "매출",
        "구매 ROAS(광고 지출 대비 수익률)": "ROAS",
        "CPC(전체) (KRW)": "CPC",
        "전환율(CVR)": "CVR",
        "CTR(전체)": "CTR",
        "클릭(전체)": "클릭",
        "동영상 재생": "동영상 재생",
        "동영상 3초 이상 재생": "동영상 3초 이상 재생",
        "동영상 100% 재생": "동영상 100% 재생",
        "보고 시작": "보고 시작",
        "보고 종료": "보고 종료"
    }
    df = df[list(column_mapping.keys())].rename(columns=column_mapping)

    # 3. 광고비 0 제거
    df = df[df["광고비"] > 0].copy()

    # 광고 게재 상태 변환 (대소문자 구분 없이 처리)
    df["상태"] = df["상태"].str.upper().map({"ACTIVE": "ON", "INACTIVE": "OFF"})

    # 평균객단가 계산 (0으로 나누기 방지)
    df["평균객단가"] = df.apply(lambda row: round(row["매출"] / row["구매"]) if row["구매"] > 0 else 0, axis=1)

    # 4. 후크 및 지속 계산 (소수점으로 계산)
    df["후크"] = (df["동영상 3초 이상 재생"] / df["동영상 재생"]).round(4)
    df["지속"] = (df["동영상 100% 재생"] / df["동영상 3초 이상 재생"]).round(4)

    # 5. 광고비 높은 순으로 정렬
    df = df.sort_values(by="광고비", ascending=False)

    # 6. 컬럼 순서 정리
    columns = ["상태", "보고 시작", "보고 종료", "제목", "광고비", "매출", "ROAS", "CPC", "CVR", "CTR", "후크", "지속", "클릭", "구매", "평균객단가"]
    df = df[columns]

    # 7. 엑셀로 저장 (임시)
    df.to_excel(output_path, index=False)

    # 8. 서식 적용
    wb = load_workbook(output_path)
    ws = wb.active
    col_idx = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    # 9. 열 너비 설정
    # 보고 시작, 보고 종료 (B, C열)
    ws.column_dimensions['A'].width = 7  # 상태
    ws.column_dimensions['B'].width = 10  # 보고 시작
    ws.column_dimensions['C'].width = 10  # 보고 종료
    
    # 제목 (D열)
    ws.column_dimensions['D'].width = 25
    
    # 광고비, 매출 (E, F열)
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    
    # ROAS, CPC, CVR, CTR, 후크, 지속, 클릭, 구매 (G-N열)
    for col in ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws.column_dimensions[col].width = 7

    # 평균객단가 (O열)
    ws.column_dimensions['O'].width = 9

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        r = row[0].row

        # 상태 변환 및 색상 적용 (A열부터 F열까지)
        try:
            status_value = ws.cell(r, col_idx["상태"]).value
            if status_value == "ON":
                status_color = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')  # 파란색
            elif status_value == "OFF":
                status_color = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid')  # 회색
            
            # 상태부터 매출까지의 열에 색상 적용
            for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:  # 상태, 보고시작, 보고종료, 제목, 광고비, 매출
                ws[f"{col_letter}{r}"].fill = status_color
        except:
            pass

        # 보고 시작/종료: 날짜 형식 (B, C열)
        for key in ["보고 시작", "보고 종료"]:
            try:
                date_value = ws.cell(r, col_idx[key]).value
                if date_value:
                    if isinstance(date_value, datetime):
                        formatted_date = date_value.strftime("%m월%d일")
                    else:
                        date_obj = datetime.strptime(str(date_value), "%Y-%m-%d")
                        formatted_date = date_obj.strftime("%m월%d일")
                    ws.cell(r, col_idx[key]).value = formatted_date
                ws.cell(r, col_idx[key]).number_format = "@"
            except:
                pass

        # ROAS: 0.00 (G열)
        ws.cell(r, col_idx["ROAS"]).number_format = "0.00"
        
        # 광고비, 매출: 원화 형식 (천 단위 구분) (E, F열)
        for key in ["광고비", "매출"]:
            try:
                val = float(ws.cell(r, col_idx[key]).value)
                ws.cell(r, col_idx[key]).number_format = "#,##0원"
            except:
                pass

        # CPC: 원화 형식 (소수점 없음) (H열)
        try:
            cpc_val = float(ws.cell(r, col_idx["CPC"]).value)
            ws.cell(r, col_idx["CPC"]).number_format = "#,##0원"
        except:
            pass

        # CVR: 보정 + 퍼센트 (I열)
        try:
            cvr_val = float(ws.cell(r, col_idx["CVR"]).value)
            if cvr_val >= 100:
                cvr_val *= 0.01
            ws.cell(r, col_idx["CVR"]).value = round(cvr_val, 4)
            ws.cell(r, col_idx["CVR"]).number_format = "0.00%"
        except:
            pass

        # CTR: 무조건 0.01 보정 후 퍼센트 (J열)
        try:
            ctr_val = float(ws.cell(r, col_idx["CTR"]).value) * 0.01
            ws.cell(r, col_idx["CTR"]).value = round(ctr_val, 4)
            ws.cell(r, col_idx["CTR"]).number_format = "0.00%"
        except:
            pass

        # 후크 / 지속: 정수 퍼센트 (K, L열)
        for key in ["후크", "지속"]:
            try:
                val = float(ws.cell(r, col_idx[key]).value)
                ws.cell(r, col_idx[key]).value = val
                ws.cell(r, col_idx[key]).number_format = "0%"
            except:
                pass

        # 클릭, 구매: 정수 (M, N열)
        for key in ["클릭", "구매"]:
            try:
                val = float(ws.cell(r, col_idx[key]).value)
                ws.cell(r, col_idx[key]).value = round(val, 0)
                ws.cell(r, col_idx[key]).number_format = "#,##0"
            except:
                pass

        # 평균객단가: 원화 형식 (O열)
        try:
            val = float(ws.cell(r, col_idx["평균객단가"]).value)
            ws.cell(r, col_idx["평균객단가"]).value = round(val, 0)
            ws.cell(r, col_idx["평균객단가"]).number_format = "#,##0원"
        except:
            pass

        # 후크 값에 따른 셀 색상 설정
        hook_value = ws.cell(r, col_idx['후크']).value
        if hook_value is not None:
            if hook_value >= 0.40:  # 40%
                fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')  # 파란색
            elif hook_value >= 0.30:  # 30%
                fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')  # 초록색
            elif hook_value >= 0.20:  # 20%
                fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')  # 주황색
            else:
                fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')  # 빨간색
        else:
            fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')  # 빨간색 (데이터 없음)
        
        ws.cell(r, col_idx['후크']).fill = fill

        # 지속 값에 따른 셀 색상 설정
        retention_value = ws.cell(r, col_idx['지속']).value
        if retention_value is not None:
            if retention_value >= 0.30:  # 30%
                fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')  # 파란색
            elif retention_value >= 0.20:  # 20%
                fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')  # 초록색
            elif retention_value >= 0.10:  # 10%
                fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')  # 주황색
            else:
                fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')  # 빨간색
        else:
            fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')  # 빨간색 (데이터 없음)
        
        ws.cell(r, col_idx['지속']).fill = fill

        # ROAS 값에 따른 셀 색상 설정
        roas_value = ws.cell(r, col_idx["ROAS"]).value
        if roas_value is not None:
            if roas_value >= 3.0:
                fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')  # 파란색
            elif roas_value >= 2.5:
                fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')  # 초록색
            elif roas_value >= 1.0:
                fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')  # 주황색
            else:
                fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')  # 빨간색
        else:
            fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')  # 빨간색 (데이터 없음)
        
        ws.cell(r, col_idx["ROAS"]).fill = fill

        # CPC 값에 따른 셀 색상 설정
        cpc_value = ws.cell(r, col_idx["CPC"]).value
        if cpc_value is not None:
            if cpc_value < 1000:
                fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')  # 파란색
            elif cpc_value < 1500:
                fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')  # 초록색
            elif cpc_value < 2000:
                fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')  # 주황색
            else:
                fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')  # 빨간색
        else:
            fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')  # 빨간색 (데이터 없음)
        
        ws.cell(r, col_idx["CPC"]).fill = fill

        # CVR 값에 따른 셀 색상 설정
        cvr_value = ws.cell(r, col_idx["CVR"]).value
        if cvr_value is not None:
            if cvr_value >= 0.07:  # 7%
                fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')  # 파란색
            elif cvr_value >= 0.05:  # 5%
                fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')  # 초록색
            elif cvr_value >= 0.03:  # 3%
                fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')  # 주황색
            else:
                fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')  # 빨간색
        else:
            fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')  # 빨간색 (데이터 없음)
        
        ws.cell(r, col_idx["CVR"]).fill = fill

        # CTR 값에 따른 셀 색상 설정
        ctr_value = ws.cell(r, col_idx['CTR']).value
        if ctr_value is not None:
            if ctr_value >= 0.05:  # 5%
                fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')  # 파란색
            elif ctr_value >= 0.03:  # 3%
                fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')  # 초록색
            elif ctr_value >= 0.02:  # 2%
                fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')  # 주황색
            else:
                fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')  # 빨간색
        else:
            fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')  # 빨간색 (데이터 없음)
        
        ws.cell(r, col_idx['CTR']).fill = fill

    wb.save(output_path)

# 테스트 실행 코드
if __name__ == "__main__":
    # 사용자에게 입력 파일 경로 물어보기
    input_file = input("변환할 Excel 파일 이름을 입력하세요: ")
    
    # 현재 디렉토리 경로 가져오기
    current_dir = os.getcwd()
    
    # 입력된 파일 경로 처리
    if not os.path.isabs(input_file):  # 상대 경로인 경우
        input_file = os.path.join(current_dir, input_file)
    
    try:
        # 데이터프레임 생성하여 날짜 정보 추출
        df = pd.read_excel(input_file)
        start_date = pd.to_datetime(df['보고 시작'].iloc[0]).strftime('%y%m%d')
        end_date = pd.to_datetime(df['보고 종료'].iloc[0]).strftime('%y%m%d')
        
        # 기본 파일명 생성 (버전 제외)
        base_filename = f"LYLYL_{start_date}_{end_date}"
        
        # 다음 버전 번호 가져오기
        version = get_next_version(base_filename)
        
        # 최종 출력 파일명 생성
        output_filename = f"{base_filename}_{version}.xlsx"
        output_file = os.path.join(current_dir, output_filename)
        
        convert_excel_file(input_file, output_file)
        print("\n✅ 변환이 완료되었습니다!")
        print(f"입력 파일: {input_file}")
        print(f"출력 파일: {output_file}")
        print("\n변환된 파일을 열어서 후크와 지속률이 올바르게 계산되었는지 확인해주세요.")
        print("- 후크: 퍼센트 값 (예: 0.75 = 75%)")
        print("- 지속: 퍼센트 값 (예: 0.5 = 50%)")
    except Exception as e:
        print(f"\n❌ 오류가 발생했습니다: {str(e)}")
        print("파일 이름을 다시 확인해주세요.")
