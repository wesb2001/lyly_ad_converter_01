import pandas as pd
from openpyxl import load_workbook
import sys
import os

def convert_excel_file(input_path: str, output_path: str):
    """
    광고 데이터 엑셀 파일을 변환하는 함수
    
    Args:
        input_path: 원본 엑셀 파일 경로
        output_path: 변환된 파일 저장 경로
    """
    print(f"파일 변환 시작: {input_path}")
    
    # 1. 파일 로드
    try:
        df = pd.read_excel(input_path)
        print(f"파일 로드 완료: 총 {len(df)} 행")
    except Exception as e:
        print(f"파일 로드 오류: {e}")
        return
    
    # 2. 컬럼 매핑
    column_mapping = {
        "광고 이름": "제목",
        "지출 금액 (KRW)": "광고비",
        "구매": "구매",
        "구매 전환값": "매출",
        "구매 ROAS(광고 지출 대비 수익률)": "ROAS",
        "CPC(전체) (KRW)": "CPC",
        "전환율(CVR)": "CVR",
        "CTR(전체)": "CTR",
        "클릭(전체)": "클릭",
        "동영상 재생": "동영상 재생",
        "동영상 3초 이상 재생": "동영상 3초 이상 재생",
        "동영상 100% 재생": "동영상 100% 재생",
    }
    
    # 필수 컬럼 확인
    missing_columns = [col for col in column_mapping.keys() if col not in df.columns]
    if missing_columns:
        print(f"경고: 다음 컬럼이 누락되었습니다: {', '.join(missing_columns)}")
        print("가능한 컬럼:", ', '.join(df.columns))
        return
    
    df = df[list(column_mapping.keys())].rename(columns=column_mapping)
    print("컬럼 매핑 완료")
    
    # 3. 광고비 0 제거
    initial_rows = len(df)
    df = df[df["광고비"] > 0].copy()
    removed_rows = initial_rows - len(df)
    print(f"광고비 0인 행 {removed_rows}개 제거됨, 남은 행: {len(df)}개")
    
    # 4. 후크 및 지속 계산 (0.01 보정 적용)
    df["후크"] = (df["동영상 3초 이상 재생"] / df["동영상 재생"]).round(4) * 0.01
    df["지속"] = (df["동영상 100% 재생"] / df["동영상 3초 이상 재생"]).round(4) * 0.01
    print("후크 및 지속 지표 계산 완료")
    
    # 5. 컬럼 순서 정리
    columns = ["제목", "광고비", "구매", "매출", "ROAS", "CPC", "CVR", "CTR", "클릭",
               "후크", "지속", "동영상 재생", "동영상 3초 이상 재생", "동영상 100% 재생"]
    df = df[columns]
    
    # 6. 엑셀로 저장 (임시)
    df.to_excel(output_path, index=False)
    print(f"기본 데이터 저장 완료: {output_path}")
    
    # 7. 서식 적용
    try:
        wb = load_workbook(output_path)
        ws = wb.active
        col_idx = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
        
        print("셀 서식 적용 중...")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            r = row[0].row
            # ROAS: 0.00
            ws.cell(r, col_idx["ROAS"]).number_format = "0.00"
            # CPC: 정수
            ws.cell(r, col_idx["CPC"]).number_format = "0"
            # CVR: 보정 + 퍼센트
            try:
                cvr_val = float(ws.cell(r, col_idx["CVR"]).value)
                if cvr_val >= 100:
                    cvr_val *= 0.01
                ws.cell(r, col_idx["CVR"]).value = round(cvr_val, 4)
                ws.cell(r, col_idx["CVR"]).number_format = "0.00%"
            except:
                pass
            # CTR: 무조건 0.01 보정 후 퍼센트
            try:
                ctr_val = float(ws.cell(r, col_idx["CTR"]).value) * 0.01
                ws.cell(r, col_idx["CTR"]).value = round(ctr_val, 4)
                ws.cell(r, col_idx["CTR"]).number_format = "0.00%"
            except:
                pass
            # 후크 / 지속: 정수 % (이미 0.01 곱해졌으므로 100 곱하지 않음)
            for key in ["후크", "지속"]:
                try:
                    val = float(ws.cell(r, col_idx[key]).value)
                    ws.cell(r, col_idx[key]).value = round(val, 4)
                    ws.cell(r, col_idx[key]).number_format = "0.00%"
                except:
                    pass
        
        wb.save(output_path)
        print("셀 서식 적용 완료")
        print(f"변환 완료! 결과가 {output_path}에 저장되었습니다.")
    except Exception as e:
        print(f"서식 적용 중 오류 발생: {e}")

# 메인 실행 부분
if __name__ == "__main__":
    # 인자가 없는 경우 사용법 안내
    if len(sys.argv) < 2:
        print("\n===== LYLYL 광고 데이터 변환기 =====")
        print("사용법:")
        print("  1. 파일 지정: python convert_excel.py 원본파일.xlsx [결과파일.xlsx]")
        print("  2. 현재 폴더 모든 파일: python convert_excel.py all")
        print("\n예시:")
        print("  python convert_excel.py LYLYL광고2025.4.14.2025.4.20.xlsx")
        print("  python convert_excel.py LYLYL광고2025.4.14.2025.4.20.xlsx 변환결과.xlsx")
        print("  python convert_excel.py all")
        sys.exit(1)
    
    # 'all' 옵션: 현재 폴더의 모든 xlsx 파일 처리
    if sys.argv[1].lower() == 'all':
        print("현재 폴더의 모든 엑셀 파일을 처리합니다...")
        current_dir = os.getcwd()
        excel_files = [f for f in os.listdir(current_dir) if f.endswith('.xlsx') and not f.endswith('_변환.xlsx')]
        
        if not excel_files:
            print("변환할 엑셀 파일을 찾을 수 없습니다.")
            sys.exit(1)
            
        for file in excel_files:
            input_path = os.path.join(current_dir, file)
            output_path = os.path.join(current_dir, file.replace('.xlsx', '_변환.xlsx'))
            print(f"\n처리 중: {file}")
            convert_excel_file(input_path, output_path)
    else:
        # 단일 파일 처리
        input_path = sys.argv[1]
        if len(sys.argv) > 2:
            output_path = sys.argv[2]
        else:
            # 출력 파일명 자동 생성
            base_name = os.path.splitext(input_path)[0]
            output_path = f"{base_name}_변환.xlsx"
            
        convert_excel_file(input_path, output_path)